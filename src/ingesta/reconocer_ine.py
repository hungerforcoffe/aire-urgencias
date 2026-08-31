"""Descarga las proyecciones de población del INE, con validación.

Qué es y por qué esta fuente
----------------------------
«Estimaciones y proyecciones de la población de Chile 2002-2035», base Censo
2017, a nivel de **comuna, sexo y edad simple**. Es el denominador que MINSAL y
el DEIS usan para calcular tasas de salud, así que usarlo deja el estudio
comparable con lo publicado.

Se prefiere a la cifra del Censo 2017 por una razón concreta: la ventana del
estudio va de 2018 a 2026 y la población cambió dentro de ella. Con un número
fijo, ese crecimiento se confunde con más urgencias.

Dos trampas de esta descarga
----------------------------
1. **La página del INE no sirve para automatizar.** Lista sus archivos con
   JavaScript, así que el HTML no contiene ni un enlace a `.xlsx`. La ruta se
   encontró probando el patrón de rutas de su gestor de contenidos.

2. **La URL que termina en `.csv` devuelve un XLSX.** El servidor ignora la
   extensión y entrega el mismo documento, con
   `Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml`.
   Es exactamente la regla 5: un 200 no es un archivo del tipo que pediste.
   Por eso aquí se valida el contenido, no la extensión.

Uso
---
    python -m src.ingesta.reconocer_ine acceso
    python -m src.ingesta.reconocer_ine descargar
"""

from __future__ import annotations

import argparse
import hashlib
import io
import logging
import sys
from pathlib import Path

import requests

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from src.rutas import LOGS, RAW, asegurar  # noqa: E402

log = logging.getLogger("ine")

BASE = ("https://www.ine.gob.cl/docs/default-source/proyecciones-de-poblacion/"
        "cuadros-estadisticos/base-2017/")
ARCHIVO = "ine_estimaciones-y-proyecciones-2002-2035_base-2017_comunas.xlsx"
DESTINO = RAW / "ine"
NOMBRE_LOCAL = "ine_proyecciones_comunas_2002-2035_base2017.xlsx"

HOJA_ESPERADA = "Est. y Proy. de Pob. Comunal"
CABECERA_ESPERADA = ["Region", "Nombre Region", "Provincia", "Nombre Provincia",
                     "Comuna", "Nombre Comuna"]
FILAS_MINIMAS = 50_000        # el archivo real trae 56.052 filas de datos
COMUNAS_CONTROL = {13101: "Santiago", 8110: "Talcahuano", 11101: "Coyhaique"}
UA = {"User-Agent": "Mozilla/5.0 (proyecto academico aire-urgencias)"}


class DescargaInvalida(Exception):
    """El archivo llegó, pero no es lo que dice ser."""


def validar(datos: bytes) -> dict:
    """Las cinco comprobaciones de la regla 5, sobre el contenido real."""
    if not datos:
        raise DescargaInvalida("0 bytes")
    if not datos.startswith(b"PK\x03\x04"):
        raise DescargaInvalida(
            f"no empieza por la firma ZIP de un xlsx; primeros bytes {datos[:16]!r}. "
            f"El INE devuelve HTML con estado 200 cuando la ruta no existe")

    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(datos), read_only=True, data_only=True)
    except Exception as e:
        raise DescargaInvalida(f"no abre como xlsx: {e}") from e

    if HOJA_ESPERADA not in wb.sheetnames:
        raise DescargaInvalida(
            f"falta la hoja {HOJA_ESPERADA!r}; tiene {wb.sheetnames}")
    ws = wb[HOJA_ESPERADA]

    filas = iter(ws.iter_rows(values_only=True))
    cabecera = [str(c).strip() if c is not None else "" for c in next(filas)]
    if cabecera[:6] != CABECERA_ESPERADA:
        raise DescargaInvalida(f"cabecera inesperada: {cabecera[:6]}")
    anios = [c for c in cabecera if c.startswith("Poblacion ")]
    if len(anios) < 30:
        raise DescargaInvalida(f"solo {len(anios)} columnas de año; se esperaban 34")

    vistas, n = set(), 0
    for fila in filas:
        n += 1
        try:
            vistas.add(int(fila[4]))
        except (TypeError, ValueError):
            pass
    if n < FILAS_MINIMAS:
        raise DescargaInvalida(f"solo {n:,} filas; se esperaban al menos {FILAS_MINIMAS:,}")

    faltan = {c: nom for c, nom in COMUNAS_CONTROL.items() if c not in vistas}
    if faltan:
        raise DescargaInvalida(
            f"no aparecen las comunas de control {faltan}: el archivo no cubre el país")

    return {"filas": n, "comunas": len(vistas), "anios": len(anios),
            "primer_anio": anios[0], "ultimo_anio": anios[-1],
            "sha256": hashlib.sha256(datos).hexdigest()}


def cmd_acceso(args) -> int:
    for ext in (".xlsx", ".csv"):
        url = BASE + ARCHIVO.replace(".xlsx", ext)
        r = requests.head(url, timeout=30, headers=UA, allow_redirects=True)
        print(f"  {r.status_code}  {ext:<6} content-type: "
              f"{r.headers.get('Content-Type', '?')[:60]}")
    print("\n  Las dos rutas devuelven el MISMO xlsx: el servidor ignora la extensión.")
    print("  Por eso la validación mira el contenido y no el nombre.")
    return 0


def cmd_descargar(args) -> int:
    url = BASE + ARCHIVO
    salida = DESTINO / NOMBRE_LOCAL
    if salida.exists() and not args.rehacer:
        print(f"  ya existe {salida.name} ({salida.stat().st_size:,} bytes); "
              f"usa --rehacer para volver a bajarlo")
        return 0

    print(f"  bajando {url[:88]}…")
    try:
        r = requests.get(url, timeout=300, headers=UA)
    except requests.RequestException as e:
        raise SystemExit(
            f"la descarga falló: {e}\n"
            f"  Si es un timeout o un 403, puede ser la IP y no el servidor: "
            f"esta conexión está detrás de CGNAT.") from e
    if r.status_code != 200:
        raise SystemExit(f"el INE devolvió {r.status_code}")

    try:
        informe = validar(r.content)
    except DescargaInvalida as e:
        cola = DESTINO / "errores"
        asegurar(cola)
        malo = cola / NOMBRE_LOCAL
        malo.write_bytes(r.content)
        raise SystemExit(
            f"  DESCARGA INVÁLIDA: {e}\n"
            f"  Guardada en {malo} para inspección. NO entra en la zona cruda.") from e

    asegurar(DESTINO)
    salida.write_bytes(r.content)
    print(f"  guardado   : {salida.relative_to(RAW.parent.parent)}")
    print(f"  tamaño     : {len(r.content):,} bytes")
    print(f"  sha256     : {informe['sha256'][:32]}…")
    print(f"  filas      : {informe['filas']:,}")
    print(f"  comunas    : {informe['comunas']}")
    print(f"  años       : {informe['anios']}  "
          f"({informe['primer_anio']} a {informe['ultimo_anio']})")
    print("  las tres comunas de control aparecen: Santiago, Talcahuano, Coyhaique")
    log.info("descarga validada: %s", informe)
    return 0


def main(argv=None):
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = p.add_subparsers(dest="cmd", required=True)
    sub.add_parser("acceso").set_defaults(fn=cmd_acceso)
    d = sub.add_parser("descargar")
    d.add_argument("--rehacer", action="store_true")
    d.set_defaults(fn=cmd_descargar)

    args = p.parse_args(argv)
    for flujo in (sys.stdout, sys.stderr):
        if hasattr(flujo, "reconfigure"):
            flujo.reconfigure(encoding="utf-8")
    asegurar(LOGS)
    logging.basicConfig(
        level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s",
        handlers=[logging.FileHandler(LOGS / "ine.log", encoding="utf-8")])
    return args.fn(args)


if __name__ == "__main__":
    raise SystemExit(main())
